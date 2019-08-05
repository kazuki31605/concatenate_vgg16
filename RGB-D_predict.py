#Cording: UTF-8

from keras.models import load_model
from keras.models import Sequential, Model
from keras.layers import Activation, Dense, Dropout
from keras.utils.np_utils import to_categorical
from keras.optimizers import Adagrad
from keras.optimizers import Adam
from keras.optimizers import SGD
from keras.initializers import TruncatedNormal, Constant
from keras.layers import Input, Dropout, Flatten, Conv2D, MaxPooling2D, Dense, Activation, BatchNormalization
from keras.preprocessing.image import array_to_img, img_to_array, load_img
import numpy as np
from PIL import Image
import os
import sys
import glob
from openpyxl import Workbook
from openpyxl.styles import PatternFill

#分類先ディレクトリ構造
#親ディレクトリの下にクラスの数だけ作り，クラス番号を1から順に割り振る
#その下に分類対象の画像を格納

label_list = []
#5クラスなので5*5で25を確保している
total = 0
ok_count = 0
num_class = 9
count = np.zeros(num_class ** 2)
ROWS = 224
COLS = 224
CHANNELS = 3

bark_img_list = []
leaf_img_list = []
form_img_list = []



result_num = 0

#結果の出力ファイル名
f = open('result_form.txt', 'w')
#モデルファイルと重みの読み込み

model = load_model('./my_model_RGB_D_10.h5')
model.load_weights('./my_model_weights_RGB_D_10.h5')

#RGB画像
test_dir = "C:/Users/hash/PycharmProjects/keras_deep/RGB-D/test_sukunai/"

for test_path in os.listdir(test_dir):
    for dir in os.listdir('{}{}'.format(test_dir, test_path)):
        print('{}{}/{}'.format(test_dir, test_path, dir))

        for file in os.listdir('{}{}/{}'.format(test_dir, test_path, dir)):
            #print('{}{}/{}'.format(test_dir, test_path, dir))
            if str(test_path) == "depth":
                label_list.append(int(dir) - 1)
                img = img_to_array(
                    load_img('{}{}/{}/{}'.format(test_dir, test_path, dir, file), target_size=(ROWS, COLS, CHANNELS)))
                bark_img_list.append(img)


            elif str(test_path) == "rgb":
                img = img_to_array(
                    load_img('{}{}/{}/{}'.format(test_dir, test_path, dir, file), target_size=(ROWS, COLS, CHANNELS)))
                leaf_img_list.append(img)



file_num = len(bark_img_list)

bark_img_list = np.array(bark_img_list)
leaf_img_list = np.array(leaf_img_list)

file_num = []

flag = 0



Y = to_categorical(label_list)


result = model.predict(([bark_img_list, leaf_img_list]))
#result_2 = model.predict_classes(([bark_img_list, form_img_list, leaf_img_list]))
#print(result)


# エクセルファイルの作成
wb = Workbook()
ws = wb.active

# 標準だと `Sheet` だが変更する場合は `title` で変更する

print(result)

j = 1

fill = PatternFill(patternType='solid', fgColor='ffff00')#最も高い色（予測ラベル)
sell = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']# クラス数のセルを用意

for value in result:
    for i, x in enumerate(value):
        #print(sell[i], i+1, x)
        if x > 0.7:
            ws[sell[i] + str(j)].fill = fill
        ws[sell[i]+str(j)] = round(x, 3)#小数点3桁で切り捨て
    j = j + 1


wb.save('./result.xlsx')
